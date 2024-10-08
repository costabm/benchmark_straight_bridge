import copy
import pandas as pd
import numpy as np
from aero_coefficients import (aero_coef, df_aero_coef_measurement_data, aero_coef_derivatives, rad, deg,
                               Cx_factor, Cy_factor, lst_methods)
from transformations import T_LsGw_func
import matplotlib.pyplot as plt
import matplotlib
import os
from my_utils import root_dir


#####################################################################################################################
# Raw Data from SOH
#####################################################################################################################
path_df = os.path.join(root_dir, r'aerodynamic_coefficients', 'aero_coef_experimental_data.csv')
df = pd.read_csv(path_df)  # raw original values
betas_uncorrected_SOH = rad(df['SOH_beta_uncorrected[deg]'].to_numpy())  # SOH initial skew angle, before performing rotation about bridge axis (which changes the beta angle).
alphas_SOH = rad(df['alpha[deg]'].to_numpy())  # Alpha: rotation about the bridge x-axis, which differs from the theta definition from L.D.Zhu and from SOH alpha (opposite direction).
betas_SOH = rad(df['beta[deg]'].to_numpy())
thetas_SOH = rad(df['theta[deg]'].to_numpy())
C_SOH_Ls = np.array([df['Cx_Ls'], df['Cy_Ls'], df['Cz_Ls'], df['Cxx_Ls'], df['Cyy_Ls'], df['Czz_Ls']])
# Adjusted Data
C_SOH_adjusted_Ls = np.array([df['Cx_Ls'], df['Cy_Ls'] * Cy_factor, df['Cz_Ls'], df['Cxx_Ls'], df['Cyy_Ls'], df['Czz_Ls']])

#####################################################################################################################
# Raw Data from CFD
#####################################################################################################################
path_df_CFD = os.path.join(root_dir, r'aerodynamic_coefficients', 'aero_coef_CFD_data.csv')
df_CFD = pd.read_csv(path_df_CFD)  # raw original values
betas_CFD = rad(df_CFD['beta[deg]'].to_numpy())
thetas_CFD = rad(df_CFD['theta[deg]'].to_numpy())
C_CFD_Ls = np.array([df_CFD['Cx_Ls'], df_CFD['Cy_Ls'], df_CFD['Cz_Ls'], df_CFD['Cxx_Ls'], df_CFD['Cyy_Ls'], df_CFD['Czz_Ls']])
# Adjusted Data
C_CFD_adjusted_Ls = np.array([df_CFD['Cx_Ls'] * Cx_factor, df_CFD['Cy_Ls'] * Cy_factor, df_CFD['Cz_Ls'], df_CFD['Cxx_Ls'], df_CFD['Cyy_Ls'], df_CFD['Czz_Ls']])

#####################################################################################################################
# Confirming some Mathematical equations
#####################################################################################################################
# thetas_yz_SOH = - alphas_SOH * np.sign(np.cos(betas_SOH))
# thetas_yz_SOH_2 = np.arcsin(np.sin(thetas_SOH)/np.sqrt(1-np.sin(betas_SOH)**2*np.cos(thetas_SOH)**2))
# thetas_SOH_2 = -np.arcsin(np.cos(betas_SOH)*np.sin(alphas_SOH))

#####################################################################################################################
# Colormaps: 2-var fittings (3D approach) VS Measurements
#####################################################################################################################
# ZOOM OUT GRAPH
# Tested Domain
beta_angle_step = 2  # in degrees.
theta_angle_step = 0.1  # in degrees.
betas = np.arange(rad(-179), rad(179)+rad(beta_angle_step)*0.012345, rad(beta_angle_step))
thetas = np.arange(rad(-10), rad(10)+rad(theta_angle_step)*0.012345, rad(theta_angle_step))
xx, yy = np.meshgrid(betas, thetas)

def colormap_2var_cons_fit_zoomout(method='2D_fit_cons', idx_to_plot=[0,1,2,3,4,5], format='original'):
    # Extrapolated coefficients
    C_Ci_grid_flat_Ls = aero_coef(xx.flatten(), yy.flatten(), method=method, coor_system='Ls')
    # Assessing the fitting at the exact SOH points, to estimate R_squared
    C_Ci_fit_at_SOH = aero_coef(betas_SOH, thetas_SOH, method=method, coor_system='Ls')
    for i in idx_to_plot:
        title_str = [r'$C_{x}$', r'$C_{y}$', r'$C_{z}$', r'$C_{rx}$', r'$C_{ry}$', r'$C_{rz}$'][i]
         # Finding the coefficient of determination R_squared
        SSres = sum((C_SOH_Ls[i] - C_Ci_fit_at_SOH[i])**2)
        SStot = sum((C_SOH_Ls[i] - np.mean(C_SOH_Ls[i])*np.ones(C_SOH_Ls[i].shape))**2)
        r_squared = 1 - SSres / SStot
        # Plotting:
        plt.figure(figsize=(5, 4), dpi=300)
        ax = plt.axes()
        plt.title('Fitting ' + title_str + ' (' + r'$R^2 = $' + "{0:.3f}".format(r_squared)+')')
        cmap = plt.get_cmap(matplotlib.cm.Spectral_r)
        absmax = max(abs(C_Ci_grid_flat_Ls[i]))
        cmap_norm = matplotlib.colors.Normalize(vmin=-absmax, vmax=absmax)
        scalarMap = matplotlib.cm.ScalarMappable(norm=cmap_norm, cmap=cmap)
        if format=='original':
            plt.scatter(xx.flatten() * 180 / np.pi, yy.flatten() * 180 / np.pi, s = np.ones(len(C_Ci_grid_flat_Ls[i]))*10, alpha=1,marker="o", c=scalarMap.to_rgba(C_Ci_grid_flat_Ls[i]))
            plt.colorbar(matplotlib.cm.ScalarMappable(norm=cmap_norm, cmap=cmap), ax=ax, alpha=1)
            plt.clim(vmin=0, vmax=0)
            ax.scatter(betas_SOH * 180 / np.pi, thetas_SOH * 180 / np.pi, s=2, color='black', label='Measurements')
            ax.set_xlabel(r'$\beta\/[\degree]$')
            ax.set_ylabel(r'$\theta\/[\degree]$')
            # handles = [handles[1], handles[0]]
            # labels = [labels[1], labels[0]]
            ax.set_xlim(deg(min(betas)), deg(max(betas)))
            plt.xticks(np.arange(-180, 180.001, 45))
            plt.yticks(np.arange(-15, 15.001, 3))
            ax.set_ylim(-90, 90)
        elif format=='TorMartin':
            plt.scatter(yy.flatten() * 180 / np.pi, xx.flatten() * 180 / np.pi, s = np.ones(len(C_Ci_grid_flat_Ls[i]))*10, alpha=1,marker="o", c=scalarMap.to_rgba(C_Ci_grid_flat_Ls[i]))
            plt.colorbar(matplotlib.cm.ScalarMappable(norm=cmap_norm, cmap=cmap), ax=ax, alpha=1)
            plt.clim(vmin=0, vmax=0)
            ax.scatter(thetas_SOH * 180 / np.pi, betas_SOH * 180 / np.pi, s=2, color='black', label='Measurements')
            ax.set_xlabel(r'$\theta\/[\degree]$')
            ax.set_ylabel(r'$\beta\/[\degree]$')
            ax.set_xlim(-10, 10)
            ax.set_xticks(np.arange(-10, 10.001, 2.5))
            ax.set_ylim(-180, 180)
            ax.set_yticks(np.arange(-180, 180.001, 30))
        handles, labels = ax.get_legend_handles_labels()
        plt.legend(handles,labels,loc=1)
        plt.tight_layout()
        plt.savefig(r'aerodynamic_coefficients/plots/3D_ZoomOut'+method+'_'+format+'_'+str(i)+'.png')
        plt.close()

# colormap_2var_cons_fit_zoomout(method='2D_fit_free', idx_to_plot=[0,1,2,3,4,5])
# colormap_2var_cons_fit_zoomout(method='2D_fit_cons', idx_to_plot=[0,1,2,3,4,5], format='TorMartin')
# colormap_2var_cons_fit_zoomout(method='2D_fit_cons', idx_to_plot=[0,1,2,3,4,5], format='TorMartin')
# colormap_2var_cons_fit_zoomout(method='2D', idx_to_plot=[1,2,3])
# colormap_2var_cons_fit_zoomout(method='cos_rule', idx_to_plot=[1,2,3])

# ZOOM IN GRAPH
# Tested Domain
beta_angle_step = 0.5  # in degrees.
theta_angle_step = 0.1  # in degrees.
betas = np.arange(rad(-5), rad(95)+rad(beta_angle_step)*0.012345, rad(beta_angle_step))
# betas = np.arange(rad(-179), rad(179)+rad(beta_angle_step)*0.012345, rad(beta_angle_step))
thetas = np.arange(rad(-10), rad(10)+rad(theta_angle_step)*0.012345, rad(theta_angle_step))
xx, yy = np.meshgrid(betas, thetas)

# def colormap_2var_cons_fit_zoomin_OLD(method='2D_fit_cons', idx_to_plot=[0,1,2,3,4,5]):
#     # Extrapolated coefficients
#     C_Ci_grid_flat_Ls = aero_coef(xx.flatten(), yy.flatten(), method=method, coor_system='Ls')
#     # Assessing the fitting at the exact SOH points, to estimate R_squared
#     C_Ci_fit_at_SOH = aero_coef(betas_SOH, thetas_SOH, method=method, coor_system='Ls')
#     for i in idx_to_plot:
#         if method == '2D':
#             title_str = [r'$C_{x}^{2D}$', r'$C_{y}^{2D}$', r'$C_{z}^{2D}$', r'$C_{rx}^{2D}$', r'$C_{ry}^{2D}$', r'$C_{rz}^{2D}$'][i]
#         elif method == 'cos_rule':
#             title_str = [r'$C_{x}^{Cosine\/rule}$', r'$C_{y}^{Cosine\/rule}$', r'$C_{z}^{Cosine\/rule}$', r'$C_{rx}^{Cosine\/rule}$', r'$C_{ry}^{Cosine\/rule}$', r'$C_{rz}^{Cosine\/rule}$'][i]
#         elif method == '2D_fit_free':
#             title_str = [r'$C_{x}^{Free}$', r'$C_{y}^{Free}$', r'$C_{z}^{Free}$', r'$C_{rx}^{Free}$', r'$C_{ry}^{Free}$', r'$C_{rz}^{Free}$'][i]
#         elif method == '2D_fit_cons':
#             title_str = [r'$C_{x}^{Constrained}$', r'$C_{y}^{Constrained}$', r'$C_{z}^{Constrained}$', r'$C_{rx}^{Constrained}$', r'$C_{ry}^{Constrained}$', r'$C_{rz}^{Constrained}$'][i]
#         # Finding the coefficient of determination R_squared
#         SSres = sum((C_SOH_Ls[i] - C_Ci_fit_at_SOH[i])**2)
#         SStot = sum((C_SOH_Ls[i] - np.mean(C_SOH_Ls[i])*np.ones(C_SOH_Ls[i].shape))**2)
#         r_squared = 1 - SSres / SStot
#         # Plotting:
#         plt.figure(figsize=(5, 4), dpi=400)
#         ax = plt.axes()
#         # plt.title('Fitting ' + title_str + ' (' + r'$R^2 = $' + "{0:.3f}".format(r_squared)+')')
#         plt.title(title_str + r'$(\beta,\theta)$ fit')
#         print(title_str+' R2 -> '+"{0:.3f}".format(r_squared))
#         cmap = plt.get_cmap(matplotlib.cm.Spectral_r)
#         absmax = max(max(abs(C_Ci_grid_flat_Ls[i])), max(abs(C_SOH_Ls[i])))
#         cmap_norm = matplotlib.colors.Normalize(vmin=-absmax, vmax=absmax)
#         scalarMap = matplotlib.cm.ScalarMappable(norm=cmap_norm, cmap=cmap)
#         plt.scatter(xx.flatten() * 180 / np.pi, yy.flatten() * 180 / np.pi, s = np.ones(len(C_Ci_grid_flat_Ls[i]))*10, alpha=1,marker="o", c=scalarMap.to_rgba(C_Ci_grid_flat_Ls[i]))
#         plt.colorbar(matplotlib.cm.ScalarMappable(norm=cmap_norm, cmap=cmap), ax=ax, alpha=1)
#         plt.clim(vmin=0, vmax=0)
#         markersize = 60
#         scatter = ax.scatter(betas_SOH * 180 / np.pi, thetas_SOH * 180 / np.pi, s=markersize, c=scalarMap.to_rgba(C_SOH_Ls[i]), label='Measurements', edgecolors='black')
#         ax.set_xlabel(r'$\beta\/[\degree]$')
#         ax.set_ylabel(r'$\theta\/[\degree]$')
#         handles, labels = ax.get_legend_handles_labels()
#         ax.set_xlim(deg(min(betas)), deg(max(betas)))
#         plt.xticks(np.arange(0, 91, 15))
#         plt.yticks(np.arange(-10, 11, 2))
#         ax.set_ylim(deg(min(thetas)), deg(max(thetas)))
#         legend = plt.legend([plt.scatter([],[], marker='o', s=markersize, edgecolors='black', facecolor=(0,0,0,0))], labels, loc=1)
#         plt.tight_layout()
#         plt.savefig(r'aerodynamic_coefficients/plots/3D_'+method+'_'+str(i)+'.png')
#         plt.close()

def colormap_2var_cons_fit_zoomin(method='2D_fit_cons', idx_to_plot=[0,1,2,3,4,5]):
    # Extrapolated coefficients
    C_Ci_grid_flat_Ls = aero_coef(xx.flatten(), yy.flatten(), method=method, coor_system='Ls')
    # Assessing the fitting at the exact SOH points, to estimate R_squared
    C_Ci_fit_at_SOH = aero_coef(betas_SOH, thetas_SOH, method=method, coor_system='Ls')
    for i in idx_to_plot:
        if method == '2D':
            title_str = [r'$C_{x}^{2D}$', r'$C_{y}^{2D}$', r'$C_{z}^{2D}$', r'$C_{rx}^{2D}$', r'$C_{ry}^{2D}$', r'$C_{rz}^{2D}$'][i]
        elif method == 'cos_rule':
            title_str = [r'$C_{x}^{Cosine\/rule}$', r'$C_{y}^{Cosine\/rule}$', r'$C_{z}^{Cosine\/rule}$', r'$C_{rx}^{Cosine\/rule}$', r'$C_{ry}^{Cosine\/rule}$', r'$C_{rz}^{Cosine\/rule}$'][i]
        elif method == '2D_fit_free':
            title_str = [r'$C_{x}^{Free}$', r'$C_{y}^{Free}$', r'$C_{z}^{Free}$', r'$C_{rx}^{Free}$', r'$C_{ry}^{Free}$', r'$C_{rz}^{Free}$'][i]
        elif method == '2D_fit_cons':
            title_str = [r'$C_{x}^{SOH}$', r'$C_{y}^{SOH}$', r'$C_{z}^{SOH}$', r'$C_{rx}^{SOH}$', r'$C_{ry}^{SOH}$', r'$C_{rz}^{SOH}$'][i]
        elif method == '2D_fit_cons_w_CFD':
            title_str = [r'$C_{x}^{SOH&CFD}$', r'$C_{y}^{SOH&CFD}$', r'$C_{z}^{SOH&CFD}$', r'$C_{rx}^{SOH&CFD}$', r'$C_{ry}^{SOH&CFD}$', r'$C_{rz}^{SOH&CFD}$'][i]
        elif method == '2D_fit_cons_scale_to_Jul':
            title_str = [r'$C_{x}^{SOH\/\/Jul.\/scaled}$', r'$C_{y}^{SOH\/\/Jul.\/scaled}$', r'$C_{z}^{SOH\/\/Jul.\/scaled}$', r'$C_{rx}^{SOH\/\/Jul.\/scaled}$', r'$C_{ry}^{SOH\/\/Jul.\/scaled}$', r'$C_{rz}^{SOH\/\/Jul.\/scaled}$'][i]
        elif method == '2D_fit_cons_w_CFD_scale_to_Jul':
            title_str = [r'$C_{x}^{SOH&CFD\/\/Jul.\/scaled}$', r'$C_{y}^{SOH&CFD\/\/Jul.\/scaled}$', r'$C_{z}^{SOH&CFD\/\/Jul.\/scaled}$', r'$C_{rx}^{SOH&CFD\/\/Jul.\/scaled}$', r'$C_{ry}^{SOH&CFD\/\/Jul.\/scaled}$', r'$C_{rz}^{SOH&CFD\/\/Jul.\/scaled}$'][i]



        # Finding the coefficient of determination R_squared
        if method == '2D_fit_cons_w_CFD_adjusted':
            SSres = sum((C_SOH_adjusted_Ls[i] - C_Ci_fit_at_SOH[i]) ** 2)
            SStot = sum((C_SOH_adjusted_Ls[i] - np.mean(C_SOH_adjusted_Ls[i]) * np.ones(C_SOH_adjusted_Ls[i].shape)) ** 2)
        else:
            SSres = sum((C_SOH_Ls[i] - C_Ci_fit_at_SOH[i])**2)
            SStot = sum((C_SOH_Ls[i] - np.mean(C_SOH_Ls[i])*np.ones(C_SOH_Ls[i].shape))**2)
        r_squared = 1 - SSres / SStot
        # Plotting:
        plt.figure(figsize=(5, 4), dpi=400)
        ax = plt.axes()
        # plt.title('Fitting ' + title_str + ' (' + r'$R^2 = $' + "{0:.3f}".format(r_squared)+')')
        # if method in ['2D', 'cos_rule']:
        plt.title(title_str + r'$(\beta,\theta)$ fit')
        # elif method is '2D_fit_free':
        #     plt.title(title_str + r'$(\beta,\theta)$ free fit. 3D depiction.')
        # elif method is '2D_fit_cons':
        #     plt.title(title_str + r'$(\beta,\theta)$ constrained fit. 3D depiction.')
        print(title_str+' R2 -> '+"{0:.3f}".format(r_squared))
        cmap = plt.get_cmap(matplotlib.cm.Spectral_r)
        absmax = max(max(abs(C_Ci_grid_flat_Ls[i])), max(abs(C_SOH_Ls[i])))
        cmap_norm = matplotlib.colors.Normalize(vmin=-absmax, vmax=absmax)
        scalarMap = matplotlib.cm.ScalarMappable(norm=cmap_norm, cmap=cmap)
        plt.scatter(xx.flatten() * 180 / np.pi, yy.flatten() * 180 / np.pi, s = np.ones(len(C_Ci_grid_flat_Ls[i]))*10, alpha=1,marker="o", c=scalarMap.to_rgba(C_Ci_grid_flat_Ls[i]))
        plt.colorbar(matplotlib.cm.ScalarMappable(norm=cmap_norm, cmap=cmap), ax=ax, alpha=1)
        plt.clim(vmin=0, vmax=0)
        markersize = 60
        if method == '2D_fit_cons_w_CFD_adjusted':
            ax.scatter(betas_SOH * 180 / np.pi, thetas_SOH * 180 / np.pi, s=markersize, c=scalarMap.to_rgba(C_SOH_adjusted_Ls[i]), label='Measurements', edgecolors='black')

        if method == '2D_fit_cons_w_CFD_scale_to_Jul':
            df = df_aero_coef_measurement_data(method)
            betas_SOHCFD, thetas_SOHCFD = rad(df['beta[deg]'].to_numpy()), rad(df['theta[deg]'].to_numpy())
            SOH_mask = df['test_case_name'].str.startswith('K71')
            CFD_mask = df['test_case_name'].str.startswith("'[b")
            C_upscaled_Ls = np.array([df['Cx_Ls'], df['Cy_Ls'], df['Cz_Ls'], df['Cxx_Ls'], df['Cyy_Ls'], df['Czz_Ls']])
            ax.scatter(betas_SOHCFD * 180 /np.pi, thetas_SOHCFD * 180 /np.pi, s=markersize, c=scalarMap.to_rgba(C_upscaled_Ls[i]), label='Measurements', edgecolors='black')
        else:
            ax.scatter(betas_SOH * 180 / np.pi, thetas_SOH * 180 / np.pi, s=markersize, c=scalarMap.to_rgba(         C_SOH_Ls[i]), label='Measurements', edgecolors='black')
        if method == '2D_fit_cons_w_CFD':
            ax.scatter(betas_CFD * 180 / np.pi, thetas_CFD * 180 / np.pi, s=markersize, c=scalarMap.to_rgba(C_CFD_Ls[i]), label='Measurements', edgecolors='black')
        elif method == '2D_fit_cons_w_CFD_adjusted':
            ax.scatter(betas_CFD * 180 / np.pi, thetas_CFD * 180 / np.pi, s=markersize, c=scalarMap.to_rgba(C_CFD_adjusted_Ls[i]), label='Measurements', edgecolors='black')
        ax.set_xlabel(r'$\beta\/[\degree]$')
        ax.set_ylabel(r'$\theta\/[\degree]$')
        handles, labels = ax.get_legend_handles_labels()
        ax.set_xlim(deg(min(betas)), deg(max(betas)))
        plt.xticks(np.arange(0, 91, 15))
        plt.yticks(np.arange(-10, 11, 2))
        ax.set_ylim(deg(min(thetas)), deg(max(thetas)))
        legend = plt.legend([plt.scatter([],[], marker='o', s=markersize, edgecolors='black', facecolor=(0,0,0,0))], labels, loc=1)
        plt.tight_layout()
        plt.savefig(r'aerodynamic_coefficients/plots/3D_'+method+'_'+str(i)+'.jpg')
        plt.close()
# colormap_2var_cons_fit_zoomin(method='2D_fit_free', idx_to_plot=[0,1,2,3,4,5])
# colormap_2var_cons_fit_zoomin(method='2D_fit_cons', idx_to_plot=[0,1,2,3,4,5])
# colormap_2var_cons_fit_zoomin(method='2D_fit_cons_scale_to_Jul', idx_to_plot=[0,1,2,3,4,5])
# colormap_2var_cons_fit_zoomin(method='2D_fit_cons_w_CFD_scale_to_Jul', idx_to_plot=[0,1,2,3,4,5])
# colormap_2var_cons_fit_zoomin(method='2D_fit_cons_2', idx_to_plot=[0,1,2,3,4,5])
# colormap_2var_cons_fit_zoomin(method='2D', idx_to_plot=[1,2,3])
# colormap_2var_cons_fit_zoomin(method='cos_rule', idx_to_plot=[1,2,3])

# def plot_2D_at_beta_fixed_OLD(method='2D_fit_cons',idx_to_plot=[0,1,2,3,4,5], plot_other_bridges=False):
#     if plot_other_bridges:
#         import os
#         import pandas as pd
#         path_raw_data = os.path.join(ROOT_DIR, r'other', 'Langenuen and other bridges - static coefficients.xlsx')
#         str_bridges = ['Langenuen', 'Julsundet', 'Sotrabru']
#         linestyles = ['dotted','dashed','dashdot']
#         df_list = [pd.read_excel(io=path_raw_data, sheet_name='Langenuen_table_only'),
#                    pd.read_excel(io=path_raw_data, sheet_name='Julsundet_table_only'),
#                    pd.read_excel(io=path_raw_data, sheet_name='Sotrabru_table_only')]
#     # ZOOM IN GRAPH
#     # Tested Domain
#     theta_angle_step = 0.1  # in degrees.
#     thetas = np.arange(rad(-10), rad(10) + rad(theta_angle_step) * 0.012345, rad(theta_angle_step))
#     beta_fixed_list = rad(np.array([0,10.1,19.9,30.1,41.2,50.9]))
#     from matplotlib.rcsetup import cycler
#     for i in idx_to_plot:
#         if method == '2D':
#             title_str = [r'$C_{x}^{2D}$', r'$C_{y}^{2D}$', r'$C_{z}^{2D}$', r'$C_{rx}^{2D}$', r'$C_{ry}^{2D}$', r'$C_{rz}^{2D}$'][i]
#         elif method == 'cos_rule':
#             title_str = [r'$C_{x}^{Cosine\/rule}$', r'$C_{y}^{Cosine\/rule}$', r'$C_{z}^{Cosine\/rule}$', r'$C_{rx}^{Cosine\/rule}$', r'$C_{ry}^{Cosine\/rule}$', r'$C_{rz}^{Cosine\/rule}$'][i]
#         elif method == '2D_fit_free':
#             title_str = [r'$C_{x}^{Free}$', r'$C_{y}^{Free}$', r'$C_{z}^{Free}$', r'$C_{rx}^{Free}$', r'$C_{ry}^{Free}$', r'$C_{rz}^{Free}$'][i]
#         elif method == '2D_fit_cons':
#             title_str = [r'$C_{x}^{Constrained}$', r'$C_{y}^{Constrained}$', r'$C_{z}^{Constrained}$', r'$C_{rx}^{Constrained}$', r'$C_{ry}^{Constrained}$', r'$C_{rz}^{Constrained}$'][i]
#         # Plotting:
#         plt.figure(figsize=(5, 4), dpi=300)
#         ax = plt.axes()
#         ax.set_prop_cycle('color',plt.cm.plasma(np.linspace(0.05,0.95,len(beta_fixed_list))))  # choos a range within colormap. e.g. [0.05,0.95] within [0,1]
#         plt.title(title_str +r'$(\beta,\theta)$ fit (section views)')
#         empty_ax = [None]*len(beta_fixed_list)
#         for b_i,beta_fixed in enumerate(beta_fixed_list):
#             marker_str = ["^","v","s","p","h","8"]
#             markersize_plt = np.array([1.8,1.8,1.8,2.3,2.3,2.3])
#             markersize_scatter = markersize_plt * 28
#             C_Ci_grid_flat_Ls = aero_coef(np.ones(len(thetas)) * beta_fixed, thetas, method=method, coor_system='Ls')
#             markevery = [0]
#             # markevery = len(thetas)-1  # could be the same as [0,-1] ?
#             plt.plot(deg(thetas), C_Ci_grid_flat_Ls[i], label=r'$\beta=$'+str(round(deg(beta_fixed),1))+'$\degree$', alpha=0.8, marker=marker_str[b_i],markevery=markevery, markersize=markersize_plt[b_i]*4, fillstyle='none')
#             measured_label = 'Measurements' if b_i == 1 else ''
#             empty_ax[b_i] = plt.scatter(deg(thetas_SOH[np.where(np.isclose(betas_SOH,beta_fixed,atol=rad(2)))]), C_SOH_Ls[i,np.where(np.isclose(betas_SOH,beta_fixed,atol=rad(2)))],alpha=0.8, s=markersize_scatter[b_i], label=measured_label,marker=marker_str[b_i], edgecolors='none')
#         ax.set_xlabel(r'$\theta\/[\degree]$')
#         ax.set_ylabel(title_str)
#         if plot_other_bridges:
#             for j in range(3):  # 3 other bridges
#                 plt.plot(df_list[j]['angle(deg)'], df_list[j]['Cy'], label=str_bridges[j]+r' ($\beta=0\degree$)', linestyle=linestyles[j], c='grey', alpha=0.7,linewidth=1)
#         handles,labels = ax.get_legend_handles_labels()
#         # handles = [handles[1], handles[0]]
#         # labels = [labels[1], labels[0]]
#         if not plot_other_bridges:
#             C_limits = [None,None,None,None,None,None]
#             plt.ylim(C_limits[i])
#         plt.xticks(np.arange(-10, 10+0.01, 2))
#         plt.tight_layout()
#         plt.savefig(r'aerodynamic_coefficients/plots/2D_beta_fixed_' + method + '_' + str(i) + '.png')
#         plt.close()
#         # Plottin legend
#     from matplotlib.legend_handler import HandlerTuple
#     if plot_other_bridges:
#         plt.figure(figsize=(2.5, 3), dpi=300)
#         plt.axis("off")
#         plt.legend(handles[:-1]+[tuple(empty_ax)],labels, handler_map={tuple: HandlerTuple(ndivide=None)})
#     else:
#         plt.figure(figsize=(2, 3), dpi=300)
#         plt.axis("off")
#         plt.legend(handles[:6]+[tuple(empty_ax)],labels, handler_map={tuple: HandlerTuple(ndivide=None)})
#     plt.tight_layout()
#     plt.savefig(r'aerodynamic_coefficients/plots/legend_2D_beta_fixed_' + method + '_' + str(i) + '.png')
#     plt.close()
#
#

def plot_2D_at_beta_fixed(method='2D_fit_cons', idx_to_plot=[0,1,2,3,4,5], plot_other_bridges=False, plot_CFD=False, plot_extra_lines=False):
    if plot_other_bridges:
        path_raw_data = os.path.join(root_dir, r'other', 'Langenuen and other bridges - static coefficients - v5.xlsx')
        str_bridges = ['Langenuen', 'Julsundet', 'Sotrabru']
        linestyles = ['dotted','dashed','dashdot']
        df_list = [pd.read_excel(io=path_raw_data, sheet_name='Langenuen_table_only'),
                   pd.read_excel(io=path_raw_data, sheet_name='Julsundet_table_only'),
                   pd.read_excel(io=path_raw_data, sheet_name='Sotrabru_table_only')]
    # ZOOM IN GRAPH
    # Tested Domain
    theta_angle_step = 0.1  # in degrees.
    thetas = np.arange(rad(-10), rad(10) + rad(theta_angle_step) * 0.012345, rad(theta_angle_step))
    beta_fixed_list = rad(np.array([0,10.1,19.9,30.1,41.2,50.9]))
    beta_extra_list = rad(np.array([60, 70, 80, 90]))
    beta_all_list = beta_fixed_list.tolist() + beta_extra_list.tolist() if plot_extra_lines else beta_fixed_list
    n_betas = len(beta_fixed_list) if not plot_extra_lines else len(beta_fixed_list)+len(beta_extra_list)
    from matplotlib.rcsetup import cycler

    for i in idx_to_plot:
        if method == '2D':
            title_str = [r'$C_{x}^{2D}$', r'$C_{y}^{2D}$', r'$C_{z}^{2D}$', r'$C_{rx}^{2D}$', r'$C_{ry}^{2D}$', r'$C_{rz}^{2D}$'][i]
        elif method == 'cos_rule':
            title_str = [r'$C_{x}^{Cosine\/rule}$', r'$C_{y}^{Cosine\/rule}$', r'$C_{z}^{Cosine\/rule}$', r'$C_{rx}^{Cosine\/rule}$', r'$C_{ry}^{Cosine\/rule}$', r'$C_{rz}^{Cosine\/rule}$'][i]
        elif method == '2D_fit_free':
            title_str = [r'$C_{x}^{Free}$', r'$C_{y}^{Free}$', r'$C_{z}^{Free}$', r'$C_{rx}^{Free}$', r'$C_{ry}^{Free}$', r'$C_{rz}^{Free}$'][i]
        elif method == '2D_fit_cons':
            title_str = [r'$C_{x}^{SOH}$', r'$C_{y}^{SOH}$', r'$C_{z}^{SOH}$', r'$C_{rx}^{SOH}$', r'$C_{ry}^{SOH}$', r'$C_{rz}^{SOH}$'][i]
        elif method == '2D_fit_cons_w_CFD':
            title_str = [r'$C_{x}^{SOH&CFD}$', r'$C_{y}^{SOH&CFD}$', r'$C_{z}^{SOH&CFD}$', r'$C_{rx}^{SOH&CFD}$', r'$C_{ry}^{SOH&CFD}$', r'$C_{rz}^{SOH&CFD}$'][i]
        elif method == '2D_fit_cons_w_CFD_adjusted':
            title_str = [r'$C_{x}^{SOH&CFD&Jul.}$', r'$C_{y}^{SOH&CFD&Jul.}$', r'$C_{z}^{SOH&CFD&Jul.}$', r'$C_{rx}^{SOH&CFD&Jul.}$', r'$C_{ry}^{SOH&CFD&Jul.}$', r'$C_{rz}^{SOH&CFD&Jul.}$'][i]
        elif method == '2D_fit_cons_scale_to_Jul':
            title_str = [r'$C_{x}^{SOH\/\/Jul.\/scaled}$', r'$C_{y}^{SOH\/\/Jul.\/scaled}$', r'$C_{z}^{SOH\/\/Jul.\/scaled}$', r'$C_{rx}^{SOH\/\/Jul.\/scaled}$', r'$C_{ry}^{SOH\/\/Jul.\/scaled}$', r'$C_{rz}^{SOH\/\/Jul.\/scaled}$'][i]
        elif method == '2D_fit_cons_w_CFD_scale_to_Jul':
            title_str = [r'$C_{x}^{SOH&CFD\/\/Jul.\/scaled}$', r'$C_{y}^{SOH&CFD\/\/Jul.\/scaled}$', r'$C_{z}^{SOH&CFD\/\/Jul.\/scaled}$', r'$C_{rx}^{SOH&CFD\/\/Jul.\/scaled}$', r'$C_{ry}^{SOH&CFD\/\/Jul.\/scaled}$', r'$C_{rz}^{SOH&CFD\/\/Jul.\/scaled}$'][i]
        elif method == '2D_fit_cons_polimi':
            title_str = [r'$C_{x}^{Polimi}$', r'$C_{y}^{Polimi}$', r'$C_{z}^{Polimi}$', r'$C_{rx}^{Polimi}$', r'$C_{ry}^{Polimi}$', r'$C_{rz}^{Polimi}$'][i]

        # Plotting:
        plt.figure(figsize=(5, 4), dpi=300)
        ax = plt.axes()
        color_list = plt.cm.plasma(np.linspace(0.05, 0.95, len(beta_fixed_list))).tolist() + plt.cm.Greys( np.linspace( 0.2, 0.95, len(beta_extra_list))).tolist()
        marker_list = ["^", "v", "s", "p", "h", "8", "D", "X", "P", "*"]
        markersize_list = np.array([1.8, 1.8, 1.8, 2.3, 2.3, 2.3, 2.3, 2.3, 2.3, 2.3]) * 28
        empty_ax_SOH = [None] * n_betas
        empty_ax_CFD = [None] * n_betas

        plt.title(title_str +r'$(\beta,\theta)$ fit ($\beta$-fixed views)')

        for b_i, beta in enumerate(beta_all_list):
            C_Ci_grid_flat_Ls = aero_coef(np.ones(len(thetas)) * beta, thetas, method=method, coor_system='Ls')
            # markevery = [0] # Alternative: markevery = len(thetas)-1  # could be the same as [0,-1] ? This is to use the marker as a visual aid as in my PhD
            # Line plots
            plt.plot(deg(thetas), C_Ci_grid_flat_Ls[i], color=color_list[b_i], label=r'$\beta=$'+str(int(round(deg(beta),0)))+'$\degree$', alpha=0.8)  # , marker=marker_str_SOH[b_i],markevery=markevery, markersize=markersize_plt[b_i]*4, fillstyle='none')
            measured_SOH_label = 'Measured' if b_i == 1 else ''
            measured_CFD_label = 'CFD' if b_i == 1 else ''
            # Scatter points
            if method == '2D_fit_cons_w_CFD_adjusted':
                empty_ax_SOH[b_i] = plt.scatter(deg(thetas_SOH[np.where(np.isclose(betas_SOH,beta, atol=rad(2)))]).tolist(), C_SOH_adjusted_Ls[i,np.where(np.isclose(betas_SOH,beta, atol=rad(2)))][0].tolist(), alpha=0.8, label=measured_SOH_label, color=color_list[b_i], marker=marker_list[b_i], s=markersize_list[b_i], edgecolors='none')

            elif method in ['2D_fit_cons_scale_to_Jul', '2D_fit_cons_w_CFD_scale_to_Jul']:
                df = df_aero_coef_measurement_data(method)
                betas_SOH_, thetas_SOH_ = rad(df['beta[deg]'].to_numpy()), rad(df['theta[deg]'].to_numpy())
                SOH_mask = df['test_case_name'].str.startswith('K71')
                CFD_mask = df['test_case_name'].str.startswith("'[b")
                C_upscaled_Ls = np.array([df['Cx_Ls'], df['Cy_Ls'], df['Cz_Ls'], df['Cxx_Ls'], df['Cyy_Ls'], df['Czz_Ls']])
                empty_ax_SOH[b_i] = plt.scatter(deg(thetas_SOH_[np.where(np.isclose(betas_SOH_, beta, atol=rad(2)) & SOH_mask)]).tolist(),     C_upscaled_Ls[i, np.where(np.isclose(betas_SOH_,beta, atol=rad(2)) & SOH_mask)][0].tolist(), alpha=0.8, label=measured_SOH_label, color=color_list[b_i], marker=marker_list[b_i], s=markersize_list[b_i], edgecolors='none')
                empty_ax_CFD[b_i] = plt.scatter(deg(thetas_SOH_[np.where(np.isclose(betas_SOH_, beta, atol=rad(2)) & CFD_mask)]).tolist(),     C_upscaled_Ls[i, np.where(np.isclose(betas_SOH_,beta, atol=rad(2)) & CFD_mask)][0].tolist(), alpha=0.8, label=measured_CFD_label, facecolor='none', marker=marker_list[b_i], s=markersize_list[b_i], edgecolors=color_list[b_i])
            else:
                empty_ax_SOH[b_i] = plt.scatter(deg(thetas_SOH[np.where(np.isclose(betas_SOH,beta, atol=rad(2)))]).tolist(),          C_SOH_Ls[i,np.where(np.isclose(betas_SOH,beta, atol=rad(2)))][0].tolist(), alpha=0.8, label=measured_SOH_label, color=color_list[b_i], marker=marker_list[b_i], s=markersize_list[b_i], edgecolors='none')
            if plot_CFD and method not in ['2D_fit_cons_scale_to_Jul', '2D_fit_cons_w_CFD_scale_to_Jul']:  # then the unfilled scatter markers shall represent CFD results, instead of being just a visual aid
                if method == '2D_fit_cons_w_CFD_adjusted':
                    empty_ax_CFD[b_i] = plt.scatter(deg(thetas_CFD[np.where(np.isclose(betas_CFD,beta, atol=rad(2)))]).tolist(), C_CFD_adjusted_Ls[i,np.where(np.isclose(betas_CFD,beta, atol=rad(2)))][0].tolist(), alpha=0.8, label=measured_CFD_label, facecolor='none', marker=marker_list[b_i], s=markersize_list[b_i], edgecolors=color_list[b_i])
                else:
                    empty_ax_CFD[b_i] = plt.scatter(deg(thetas_CFD[np.where(np.isclose(betas_CFD,beta, atol=rad(2)))]).tolist(),          C_CFD_Ls[i,np.where(np.isclose(betas_CFD,beta, atol=rad(2)))][0].tolist(), alpha=0.8, label=measured_CFD_label, facecolor='none', marker=marker_list[b_i], s=markersize_list[b_i], edgecolors=color_list[b_i])
        ax.set_xlabel(r'$\theta\/[\degree]$')
        y_label_str = [r'$C_{x}$', r'$C_{y}$', r'$C_{z}$', r'$C_{rx}$', r'$C_{ry}$', r'$C_{rz}$'][i]
        ax.set_ylabel(y_label_str)
        if plot_other_bridges:
            other_bridge_coef_strings = ['Cx', 'Cy', 'Cz', 'Crx', 'Cry', 'Crz']
            for j in [1]:  # 1, 2, or 3 other bridges
                try:
                    ax.plot(df_list[j]['angle(deg)'], df_list[j][other_bridge_coef_strings[i]], label=str_bridges[j]+r' ($\beta=0\degree$)', linestyle=linestyles[j], color=color_list[0], alpha=0.7,linewidth=1)
                except:
                    pass
        handles, labels = ax.get_legend_handles_labels()
        if plot_CFD:
            handles[0], handles[1], handles[2], handles[3] = handles[2], handles[3], handles[0], handles[1]  # swapping
            labels[0], labels[1], labels[2], labels[3] = labels[2], labels[3], labels[0], labels[1]  # swapping
        else:
            handles[0], handles[1], handles[2] = handles[2], handles[0], handles[1]  # swapping
            labels[0], labels[1], labels[2] = labels[2], labels[0], labels[1]  # swapping


        # handles = [handles0[2], handles0[3], handles0[0], handles0[1], handles0[4], handles0[5], handles0[6], handles0[7], handles0[8], handles0[9], handles0[10], handles0[11], handles0[12]]
        # labels = [labels0[2], labels0[3], labels0[0], labels0[1], labels0[4], labels0[5], labels0[6], labels0[7], labels0[8], labels0[9], labels0[10], labels0[11], labels0[12]]
        # if plot_other_bridges:
        #     handles += handles0[13]
        #     labels += labels0[13]
        if not plot_other_bridges:
            C_limits = [None,None,None,None,None,None]
            plt.ylim(C_limits[i])
        plt.xticks(np.arange(-10, 10+0.01, 2))
        # if plot_extra_lines:
        #     # ylims = [[-0.05,0.002], [-0.04,0.08], [-0.8,0.7], [-0.2,0.2], [None,None], [None,None]]
        #     ylims = [[-0.05, 0.002], [-0.04, 0.09], [-0.8, 0.7], [-0.2, 0.2], [None, None], [None, None]]
        #     if i in [0,1,2,3]:
        #         plt.ylim(ylims[i])
        ylims = [[-0.046, 0.002], [-0.04, 0.11], [-1.0, 0.6], [-0.16, 0.21], [None, None], [None, None]]
        plt.ylim(ylims[i])
        plt.tight_layout()
        plt.savefig(os.path.join(root_dir, r'aerodynamic_coefficients/plots/2D_beta_fixed_' + method + '_' + str(i) + '.jpg'))
        plt.close()

        # Plotting legend
        if i == 1:
            from matplotlib.legend_handler import HandlerTuple
            if plot_other_bridges:
                plt.figure(figsize=(2.3, 3.2), dpi=1000)
                plt.axis("off")
                if plot_CFD:
                    plt.legend([tuple(empty_ax_SOH)] + [tuple(empty_ax_CFD)] + handles[2:], labels, handler_map={tuple: HandlerTuple(ndivide=None)})
                else:
                    plt.legend([tuple(empty_ax_SOH)] + handles[1:], labels, handler_map={tuple: HandlerTuple(ndivide=None)})
            else:
                plt.figure(figsize=(2, 3), dpi=1000)
                plt.axis("off")
                plt.legend(handles, labels) #, handler_map={tuple: HandlerTuple(ndivide=None)})
            plt.tight_layout()
            plt.savefig(r'aerodynamic_coefficients/plots/legend_2D_beta_fixed_' + method + '.jpg')
            plt.close()

# plot_2D_at_beta_fixed(method='2D_fit_free', idx_to_plot=[0,1,2,3,4,5], plot_other_bridges=False)
# plot_2D_at_beta_fixed(method='2D_fit_cons', idx_to_plot=[0,1,2,3,4,5], plot_other_bridges=True, plot_CFD=False, plot_extra_lines=True)
# plot_2D_at_beta_fixed(method='2D_fit_cons_scale_to_Jul', idx_to_plot=[0,1,2,3,4,5], plot_other_bridges=True, plot_CFD=False, plot_extra_lines=True)
# plot_2D_at_beta_fixed(method='2D_fit_cons_w_CFD_scale_to_Jul', idx_to_plot=[3], plot_other_bridges=True, plot_CFD=False, plot_extra_lines=True)
# plot_2D_at_beta_fixed(method='2D_fit_cons_w_CFD', idx_to_plot=[0,1,2,3,4,5], plot_other_bridges=True, plot_CFD=True,  plot_extra_lines=True)
# plot_2D_at_beta_fixed(method='2D_fit_cons_w_CFD_adjusted', idx_to_plot=[0,1,2,3,4,5], plot_other_bridges=True, plot_CFD=True,  plot_extra_lines=True)
# plot_2D_at_beta_fixed(method='2D_fit_cons_2', idx_to_plot=[0,1,2,3,4,5], plot_other_bridges=False)
# plot_2D_at_beta_fixed(method='2D', idx_to_plot=[1,2,3], plot_other_bridges=False)
# plot_2D_at_beta_fixed(method='cos_rule', idx_to_plot=[1,2,3], plot_other_bridges=False)

def plot_2D_at_beta_fixed_polimi(method='2D_fit_cons_polimi', idx_to_plot=[0, 1, 2, 3, 4, 5], deg_list=None, zoom='in',
                                 beta_list=rad(np.array([0, 10, 20, 30, 40, 50, 60, 70, 80, 90])), coor_system='Ls'):
    if method[-5:] == '.xlsx':
        method_is_table = True
        coor_system = (method.split('aero_coefs_')[1]).split('_', maxsplit=1)[0]
        method_behind_table = ((method.split('aero_coefs_')[1]).split('_', maxsplit=1)[1]).split('.xlsx')[0]  # obtains the string after aero_coefs_Ls_ or aero_coefs_Gw_ or aero_coefs_Lnw_ and before .xlsx
        assert deg_list is None
    else:
        method_is_table = False
        if deg_list is None:
            deg_list = aero_coef.__defaults__[0][method]  # retrieving the default parameter values :)
    # ZOOM IN GRAPH
    # Tested Domain
    if zoom == 'in':
        theta_angle_step = 0.1  # in degrees.
        thetas = np.arange(rad(-10), rad(10) + rad(theta_angle_step) * 0.012345, rad(theta_angle_step))
        xlims = [-10.4, 10.4]
    elif zoom == 'inin':
        theta_angle_step = 0.1  # in degrees.
        thetas = np.arange(rad(-2.5), rad(2.5) + rad(theta_angle_step) * 0.012345, rad(theta_angle_step))
        xlims = [-2.5, 2.5]
    elif zoom == 'out':
        theta_angle_step = 1  # in degrees.
        thetas = np.arange(rad(-89.99), rad(89.99) + rad(theta_angle_step) * 0.012345, rad(theta_angle_step))
        xlims = [-91, 91]
    n_betas = len(beta_list)
    from matplotlib.rcsetup import cycler

    for i in idx_to_plot:
        assert method in lst_methods
        if coor_system == 'Ls':
            title_str = [r'$C_{x}^{Polimi}$', r'$C_{y}^{Polimi}$', r'$C_{z}^{Polimi}$', r'$C_{rx}^{Polimi}$', r'$C_{ry}^{Polimi}$', r'$C_{rz}^{Polimi}$'][i]
        elif coor_system == 'Gw':
            title_str = [r'$C_{Xu}^{Polimi}$', r'$C_{Yv}^{Polimi}$', r'$C_{Zw}^{Polimi}$', r'$C_{rXu}^{Polimi}$', r'$C_{rYv}^{Polimi}$', r'$C_{rZw}^{Polimi}$'][i]

        # Plotting:
        plt.figure(figsize=(5, 4), dpi=300)
        ax = plt.axes()
        # color_list = plt.cm.plasma(np.linspace(0, 0.95, len(beta_list))).tolist()
        color_list = plt.cm.turbo(np.linspace(0, 0.95, len(beta_list))).tolist()
        marker_list = ["^", "v", "s", "D", "p", "h", "8", "X", "P", "*"]
        markersize_list = np.array([1.8, 1.8, 1.8, 2.3, 2.3, 2.3, 2.3, 2.3, 2.3, 2.3]) * 28
        empty_ax_polimi = [None] * n_betas

        plt.title(title_str +r'$(\beta,\theta)$ fit ($\beta$-fixed views)')

        for b_i, beta in enumerate(beta_list):
            C_Ci_grid_flat = aero_coef(np.ones(len(thetas)) * beta, thetas, method=method, coor_system=coor_system,
                                          degree_list={method: deg_list})
            # Line plots
            plt.plot(deg(thetas), C_Ci_grid_flat[i], color=color_list[b_i], label=r'$\beta=$'+str(int(round(deg(beta),0)))+'$\degree$', alpha=0.8)  # , marker=marker_str_polimi[b_i],markevery=markevery, markersize=markersize_plt[b_i]*4, fillstyle='none')
            measured_label = 'Measured' if b_i == 1 else ''

            if '_polimi' in method:
                if method_is_table:  # then the method is table. But we want to plot the raw wind tunnel data, so we go get it from the method_behind_table which was previously used to arrive at the table at hand.
                    df = df_aero_coef_measurement_data(method_behind_table)
                else:
                    df = df_aero_coef_measurement_data(method)
                betas_polimi, thetas_polimi = rad(df['beta[deg]'].to_numpy()), rad(df['theta[deg]'].to_numpy())
                C_upscaled_Ls = np.array([df['Cx_Ls'], df['Cy_Ls'], df['Cz_Ls'], df['Cxx_Ls'], df['Cyy_Ls'], df['Czz_Ls']])
                if coor_system == 'Ls':
                    C_upscaled = C_upscaled_Ls.copy()
                elif coor_system == 'Gw':
                    T_GwLs = np.transpose(T_LsGw_func(betas_polimi, thetas_polimi, dim='6x6'), axes=(0, 2, 1))
                    C_upscaled_Gw = np.einsum('nij,jn->in', T_GwLs, C_upscaled_Ls, optimize=True)
                    C_upscaled = C_upscaled_Gw.copy()
                thetas_to_plot = deg(thetas_polimi[np.where(np.isclose(betas_polimi, beta, atol=rad(2)))]).tolist()
                C_upscaled_to_plot = C_upscaled[i, np.where(np.isclose(betas_polimi, beta, atol=rad(2)))][0].tolist()
                empty_ax_polimi[b_i] = plt.scatter(thetas_to_plot, C_upscaled_to_plot, alpha=0.8, label=measured_label, color=color_list[b_i], marker=marker_list[b_i], s=markersize_list[b_i], edgecolors='none')
        ax.set_xlabel(r'$\theta\/[\degree]$')
        if coor_system == 'Ls':
            y_label_str = [r'$C_{x}$', r'$C_{y}$', r'$C_{z}$', r'$C_{rx}$', r'$C_{ry}$', r'$C_{rz}$'][i]
        elif coor_system == 'Gw':
            y_label_str = [r'$C_{Xu}$', r'$C_{Yv}$', r'$C_{Zw}$', r'$C_{rXu}$', r'$C_{rYv}$', r'$C_{rZw}$'][i]
        ax.set_ylabel(y_label_str)
        handles, labels = ax.get_legend_handles_labels()
        if len(beta_list) > 3:
            handles.append(handles.pop(2))  # replacing the 3rd item ("Measured") to last
            labels.append(labels.pop(2))  # replacing the 3rd item ("Measured") to last

        C_limits = [None,None,None,None,None,None]
        plt.ylim(C_limits[i])
        if zoom == 'in':
            plt.xticks(np.arange(-10, 10+0.01, 2))
        if zoom == 'inin':
            plt.xticks(np.arange(-3, 3 + 0.01, 1))
        if '2D_fit_cons_polimi-K12-G-L-TS-SVV' in method and coor_system == 'Ls':
            if zoom == 'in':
                ylims = [[-0.025, 0.001], [-0.04, 0.125], [-1.07, 0.60], [-0.165, 0.235], [None, None], [None, None]]
            elif zoom == 'inin':
                ylims = [[-0.025, 0.001], [-0.005, 0.125], [-0.45, 0.18], [-0.08, 0.04], [None, None], [None, None]]
            plt.ylim(ylims[i])
        plt.xlim(xlims)
        plt.grid()
        plt.tight_layout()
        # File name
        fig_path_name = r'aerodynamic_coefficients/plots/2D_beta_fixed_'
        if len(beta_list) == 1:
            assert beta_list[0] == 0
            fig_path_name += 'at_0_'
        if method_is_table:
            fig_path_name += 'xlsx_' + method.split('.xlsx')[0] + "_" + coor_system + "_" + zoom + "_" + "_dof_" + str(i) + '.jpg'
        else:
            fig_path_name += method + "_" + coor_system + "_" + zoom + "_" + '_dof_' + str(i) + f'_deg_{deg_list[i]}' + '.jpg'

        plt.savefig(os.path.join(root_dir, fig_path_name))
        plt.close()

        # Plotting legend
        legend_path_name = fig_path_name.split(r'2D_beta_fixed_')[0] + 'legend_2D_beta_fixed_' + fig_path_name.split(r'2D_beta_fixed_')[1]
        if i == 1:
            from matplotlib.legend_handler import HandlerTuple
            plt.figure(figsize=(10, 0.8), dpi=1000)
            plt.axis("off")
            plt.legend(handles, labels, ncol=6) #, handler_map={tuple: HandlerTuple(ndivide=None)})
            plt.tight_layout()
            plt.savefig(os.path.join(root_dir, legend_path_name))
            plt.close()

        # # If you just want to plot the markers, uncomment this:
        # plt.figure(figsize=(1, 3))
        # color_list = plt.cm.turbo(np.linspace(0, 0.95, 10)).tolist()
        # marker_list = ["^", "v", "s", "D", "p", "h", "8", "X", "P", "*"]
        # markersize_list = np.array([1.8, 1.8, 1.8, 2.3, 2.3, 2.3, 2.3, 2.3, 2.3, 2.3]) * 28
        # [plt.scatter(i, 0, c=color_list[i], marker=marker_list[i], s=markersize_list[i]) for i in range(10)]
        # plt.savefig(os.path.join(root_dir, "measured_markers.jpg"))


# plot_2D_at_beta_fixed_polimi(method='aero_coefs_Gw_2D_fit_cons_polimi-K12-G-L-TS-SVV.xlsx', idx_to_plot=[0,1,2,3,4,5], deg_list=None, zoom='in', beta_list=rad(np.array([0])))
# plot_2D_at_beta_fixed_polimi(method='aero_coefs_Ls_2D_fit_cons_polimi-K12-G-L-TS-SVV.xlsx', idx_to_plot=[0,1,2,3], deg_list=None, zoom='in', beta_list=rad(np.array([0])))
plot_2D_at_beta_fixed_polimi(method='aero_coefs_Ls_2D_fit_cons_polimi-K12-G-L-TS-SVV.xlsx', idx_to_plot=[0,1,2,3], deg_list=None, zoom='in')
plot_2D_at_beta_fixed_polimi(method='cos_rule_aero_coefs_Ls_2D_fit_cons_polimi-K12-G-L-TS-SVV.xlsx', idx_to_plot=[0,1,2,3], deg_list=None, zoom='in')
# plot_2D_at_beta_fixed_polimi(method='aero_coefs_Ls_2D_fit_cons_polimi-K12-G-L-TS-SVV.xlsx', idx_to_plot=[0,1,2,3], deg_list=None, zoom='inin')
# plot_2D_at_beta_fixed_polimi(method='cos_rule_aero_coefs_Ls_2D_fit_cons_polimi-K12-G-L-TS-SVV.xlsx', idx_to_plot=[0,1,2,3], deg_list=None, zoom='inin')



# for d in [2]: #,3,4,5,6,7,8,9]:
    # plot_2D_at_beta_fixed_polimi(method='2D_fit_cons_polimi-K12-G-L-TS-SVV', idx_to_plot=[1], deg_list=[d,d,d,d,d,d], zoom='in')
    # plot_2D_at_beta_fixed_polimi(method='2D_fit_cons_polimi-K12-G-L-T1-SVV', idx_to_plot=[0,1,2,3], deg_list=[d,d,d,d,d,d], zoom='in')
    # plot_2D_at_beta_fixed_polimi(method='2D_fit_cons_polimi-K12-G-L-T3-SVV', idx_to_plot=[0,1,2,3], deg_list=[d,d,d,d,d,d], zoom='in')
    # plot_2D_at_beta_fixed_polimi(method='2D_fit_cons_polimi-K12-G-L-CS-SVV', idx_to_plot=[0,1,2,3], deg_list=[d,d,d,d,d,d], zoom='in')
    # plot_2D_at_beta_fixed_polimi(method='2D_fit_cons_polimi-K12-G-L-SVV', idx_to_plot=[0,1,2,3], deg_list=[d,d,d,d,d,d], zoom='in')
    # plot_2D_at_beta_fixed_polimi(method='2D_fit_free_polimi', idx_to_plot=[1,2,3], deg_list=[d,d,d,d,d,d], zoom='in')


def table_r_squared_polimi(deg_min=2, deg_max=9, method='2D_fit_cons_polimi', export_table=True):
    #####################################################################################################################
    # Raw Data from Polimi
    #####################################################################################################################
    df = df_aero_coef_measurement_data(method)
    betas_polimi = rad(df['beta[deg]'].to_numpy())
    thetas_polimi = rad(df['theta[deg]'].to_numpy())
    C_polimi_Ls = np.array([df['Cx_Ls'], df['Cy_Ls'], df['Cz_Ls'], df['Cxx_Ls'], df['Cyy_Ls'], df['Czz_Ls']])

    table_r_squared = np.empty((len(range(deg_min, deg_max+1)), 7)) * np.nan
    for j, d in enumerate(range(deg_min, deg_max+1)):
        fit_degree_list = [d, d, d, d, d, d]
        C_Ci_fit_at_polimi = aero_coef(betas_polimi, thetas_polimi, method=method, coor_system='Ls',
                                       degree_list={method: fit_degree_list})
        table_r_squared[j, 0] = d
        for i in range(6):
            # Finding the coefficient of determination R_squared
            SSres = sum((C_polimi_Ls[i] - C_Ci_fit_at_polimi[i]) ** 2)
            SStot = sum((C_polimi_Ls[i] - np.mean(C_polimi_Ls[i]) * np.ones(C_polimi_Ls[i].shape)) ** 2)
            table_r_squared[j, i+1] = 1 - SSres / SStot
    df_table_r_squared = pd.DataFrame(table_r_squared, columns=['poly_degree', 'R2_Cx', 'R2_Cy', 'R2_Cz', 'R2_Crx',
                                                                'R2_Cry', 'R2_Crz',])
    if export_table:
        df_table_r_squared.to_csv(os.path.join(root_dir, r'aerodynamic_coefficients/plots/table_r_squared_'+method+'.csv'))
    return table_r_squared, df_table_r_squared
# table_r_squared, df_table_r_squared = table_r_squared_polimi(method='2D_fit_cons_polimi-K12-G-L-TS-SVV')
# table_r_squared, df_table_r_squared = table_r_squared_polimi(method='2D_fit_free_polimi')


raise NotImplementedError


def plot_2D_at_theta_0(idx_to_plot=[0,1,2,3,4,5], plot_for_EACWE2022=False):
    beta_angle_step = 0.1  # in degrees.
    betas = np.arange(rad(-15), rad(105) + rad(theta_angle_step) * 0.012345, rad(beta_angle_step))
    # thetas = np.zeros(betas.shape)
    thetas = np.zeros(betas.shape) # * rad(0)
    # C_Ci_grid_flat_Ls_2D_cos        = aero_coef(betas, thetas, method='cos_rule'           , coor_system='Ls')
    C_Ci_grid_flat_Ls_2D            = aero_coef(betas, thetas, method='2D'           , coor_system='Ls')
    C_Ci_grid_flat_Ls_2D_fit_free   = aero_coef(betas, thetas, method='2D_fit_free'  , coor_system = 'Ls')
    C_Ci_grid_flat_Ls_2D_fit_cons   = aero_coef(betas, thetas, method='2D_fit_cons'  , coor_system = 'Ls')
    # C_Ci_grid_flat_Ls_2D_fit_cons_2 = aero_coef(betas, thetas, method='2D_fit_cons_2', coor_system = 'Ls')
    for i in idx_to_plot:
        plt.figure(figsize=(5, 4), dpi=300)
        title_str = [r'$C_{x}$', r'$C_{y}$', r'$C_{z}$', r'$C_{rx}$', r'$C_{ry}$', r'$C_{rz}$'][i]
        # Plotting:
        plt.title(title_str + r'$(\beta,\theta)$ fit at $\theta=0\degree$')
        plt.xlabel(r'$\beta\/[\degree]$')
        plt.plot(deg(betas),C_Ci_grid_flat_Ls_2D_fit_cons[i], label='3D' if plot_for_EACWE2022 else r'Constr. fit', linewidth=2, alpha=0.8, ls='-', color='brown')
        # plt.plot(deg(betas),C_Ci_grid_flat_Ls_2D_fit_cons_2[i], label=r'2-var. constr. fit (2)', linewidth=2, alpha=0.7, ls='--', color='brown')
        # plt.plot(deg(betas),C_Ci_grid_flat_Ls_2D_cos[i], label='Cosine rule', linewidth=2., alpha=0.8, ls='--', color='gold')
        plt.plot(deg(betas),C_Ci_grid_flat_Ls_2D[i], label='Cos. rule' if plot_for_EACWE2022 else '2D approach', linewidth=2., alpha=0.8, ls='-' if plot_for_EACWE2022 else '--', color='green' if plot_for_EACWE2022 else 'gold')
        for SOH_idx, b, t in zip(range(len(betas_SOH)), betas_SOH, thetas_SOH):
            if plot_for_EACWE2022:
                if t == 0 and b == 0:
                    plt.scatter(deg(b), C_SOH_Ls[i,SOH_idx], alpha=0.9, s=30, label=r'Measur. '+r'$(\beta=0\degree)$', marker="o", c='black')
                elif t == 0:
                    label = r'Measur. '+r'$(\beta\neq0\degree)$' if SOH_idx==7 else None
                    plt.scatter(deg(b), C_SOH_Ls[i, SOH_idx], alpha=0.9, s=30, label=label, marker="x", c='black')
                plt.xlim([-2,92])
                plt.xticks(np.arange(0, 91, 15))
            else:
                if t == 0:
                    marker_str = "o"
                    alpha = 0.9
                    color = 'black'
                    if b == 0:
                        label=r'Measur. '+r'$(\theta=0\degree)$'
                else:
                    marker_str = "x"
                    alpha = 0.7
                    color = 'grey'
                    label = None
                    if SOH_idx ==0 :
                        label=r'Measur. '+r'$(\theta\neq0\degree)$'
                plt.scatter(deg(b), C_SOH_Ls[i,SOH_idx], alpha=alpha, s=20, label=label, marker=marker_str, c=color) #edgecolors='none'
                plt.xlim([-15,105])
                plt.xticks(np.arange(-15, 106, 15))
        if i in [4,5]:  # for Cry and Crz
            plt.ylim(plt.gca().get_ylim())  # the free polynomial fitting shall not be inside the visible graph
        None if plot_for_EACWE2022 else plt.plot(deg(betas),C_Ci_grid_flat_Ls_2D_fit_free[i], label=r'Free fit', linewidth=2, alpha=0.8, ls='-.', color='green') # ls=(0, (3, 1.5, 1, 1.5, 1, 1.5))
        plt.tight_layout()
        handles,labels = plt.gca().get_legend_handles_labels()
        if plot_for_EACWE2022:
            plt.savefig(r'aerodynamic_coefficients/plots/EACWE22_2D_theta_0_'+str(i)+'.png')
        else:
            plt.savefig(r'aerodynamic_coefficients/plots/2D_theta_0_'+str(i)+'.png')
        plt.close()
    # Plotting legend
    plt.figure(figsize=(6, 3), dpi=300)
    plt.axis("off")
    from matplotlib.legend_handler import HandlerTuple
    if not plot_for_EACWE2022:
        handles = [handles[1], handles[2], handles[0], handles[3], handles[4]]
        labels = [labels[1], labels[2], labels[0], labels[3], labels[4]]
    plt.legend(handles, labels, handler_map={tuple: HandlerTuple(ndivide=None)}, ncol=2)
    plt.tight_layout()
    plt.savefig(r'aerodynamic_coefficients/plots/2D_theta_0_legend.png')
    plt.close()

    # Plotting the derivative of Cz
    _, C_Ci_dtheta_grid_flat_Ls_cos_rule = aero_coef_derivatives(betas, thetas, method='cos_rule'           , coor_system='Ls')
    _, C_Ci_dtheta_grid_flat_Ls_2D = aero_coef_derivatives(betas, thetas, method='2D', coor_system='Ls')
    _, C_Ci_dtheta_grid_flat_Ls_2D_fit_free = aero_coef_derivatives(betas, thetas, method='2D_fit_free'  , coor_system='Ls')
    _, C_Ci_dtheta_grid_flat_Ls_2D_fit_cons = aero_coef_derivatives(betas, thetas, method='2D_fit_cons'  , coor_system='Ls')
    C_SOH_Ls_reshape = np.moveaxis(C_SOH_Ls.reshape(6,6,5), -1,-2)  # same shape as in Skew wind Paper 2
    betas_SOH_reshape = np.moveaxis(betas_SOH.reshape(6, 5), -1, -2)
    thetas_SOH_reshape = np.moveaxis(thetas_SOH.reshape(6,5), -1,-2)
    C_SOH_Ls_dtheta = np.moveaxis(np.array([[np.gradient(C_SOH_Ls_reshape[c,:,b], thetas_SOH_reshape[:,b], axis=0, edge_order=1) for b in range(6)] for c in range(6)]), -1,-2)
    i = 2
    plt.figure(figsize=(5, 4), dpi=300)
    title_str = [r'$\partial$$C_{x}$', r'$\partial$$C_{y}$', r'$\partial$$C_{z}$', r'$\partial$$C_{rx}$', r'$\partial$$C_{ry}$', r'$\partial$$C_{rz}$'][idx_to_plot]
    # Plotting:
    plt.title(title_str + r'$/ \partial\theta $ at $\theta=0\degree$')
    plt.xlabel(r'$\beta\/[\degree]$')
    plt.plot(deg(betas), C_Ci_dtheta_grid_flat_Ls_cos_rule[i], label='Cos. rule', linewidth=2., alpha=0.8, ls='-', color='green')
    # plt.plot(deg(betas), C_Ci_dtheta_grid_flat_Ls_2D[i], label='2D', linewidth=2., alpha=0.8, ls='-', color='yellow')
    plt.plot(deg(betas), C_Ci_dtheta_grid_flat_Ls_2D_fit_cons[i], label='3D', linewidth=2, alpha=0.8, ls='-', color='brown')
    # plt.plot(deg(betas), C_Ci_dtheta_grid_flat_Ls_2D_fit_free[i], label='3D free', linewidth=2, alpha=0.8, ls='-', color='pink')
    marker_str = "o"
    alpha = 0.9
    color = 'black'
    label=r'Measur. '+r'$(\theta=0\degree)$'
    plt.scatter(deg(betas_SOH_reshape[2,:].flatten()), C_SOH_Ls_dtheta[i,2,:].flatten(), alpha=alpha, s=20, label=label, marker=marker_str, c=color) #edgecolors='none'
    plt.xlim([-15,105])
    if i in [4,5]:  # for Cry and Crz
        plt.ylim(plt.gca().get_ylim())  # the free polynomial fitting shall not be inside the visible graph
    None if plot_for_EACWE2022 else plt.plot(deg(betas),C_Ci_grid_flat_Ls_2D_fit_free[i], label=r'Free fit', linewidth=2, alpha=0.8, ls='-.', color='green') # ls=(0, (3, 1.5, 1, 1.5, 1, 1.5))
    plt.tight_layout()
    plt.xticks(np.arange(-15,106,15))
    handles,labels = plt.gca().get_legend_handles_labels()
    plt.savefig(r'aerodynamic_coefficients/plots/2D_theta_0_dtheta_'+str(i)+'.png')
    # plt.savefig(r'aerodynamic_coefficients/plots/2D_theta_0_'+method+'_'+str(i)+'.png')
    plt.show()

plot_2D_at_theta_0(idx_to_plot=[0,1,2,3,4,5])

def table_r_squared_func(deg_min=1, deg_max=6, method='2D_fit_cons', export_table=True):
    table_r_squared = np.empty((len(range(deg_min,deg_max+1)),7)) * np.nan
    for j, d in enumerate(range(deg_min, deg_max+1)):
        fit_degree_list = [d, d, d, d, d, d]
        C_Ci_fit_at_SOH = aero_coef(betas_SOH, thetas_SOH, method=method, coor_system='Ls',
                                    constr_fit_degree_list=fit_degree_list, constr_fit_2_degree_list=fit_degree_list,
                                    free_fit_degree_list=fit_degree_list)
        table_r_squared[j,0] = d
        for i in range(6):
            # Finding the coefficient of determination R_squared
            SSres = sum((C_SOH_Ls[i] - C_Ci_fit_at_SOH[i]) ** 2)
            SStot = sum((C_SOH_Ls[i] - np.mean(C_SOH_Ls[i]) * np.ones(C_SOH_Ls[i].shape)) ** 2)
            table_r_squared[j, i+1] = 1 - SSres / SStot
    df_table_r_squared = pd.DataFrame(table_r_squared, columns=['poly_degree', 'R2_Cx', 'R2_Cy', 'R2_Cz', 'R2_Crx',
                                                                'R2_Cry', 'R2_Crz',])
    if export_table: df_table_r_squared.to_csv(r'aerodynamic_coefficients/plots/table_r_squared_'+method+'.csv')
    return table_r_squared, df_table_r_squared
# table_r_squared, df_table_r_squared = table_r_squared_func(method='2D_fit_free')
# table_r_squared, df_table_r_squared = table_r_squared_func(method='2D_fit_cons')
# table_r_squared, df_table_r_squared = table_r_squared_func(method='2D_fit_cons_2')
# table_r_squared, df_table_r_squared = table_r_squared_func(method='2D')
# table_r_squared, df_table_r_squared = table_r_squared_func(method='cos_rule')

def color_table_SOH_values(color_interval='same_as_zoomin_plot', color_in='font', method='2D_fit_cons'):
    """
    Colored table with the 30 experimental results for each aero coefficient
    :param color_interval: 'same_as_zoomin_plot' -> each cell has same color as the experimental "circles" in the 2D plots
    :param color_in: 'font' or 'cell' -> either paint the cell or the font
    :param method: Only for color normalization purposes
    :param fit_degree_list: Only for color normalization purposes
    :return:
    """
    # ########################################################
    # Small tables format, to publish in the paper
    filename = r"aerodynamic_coefficients\\aero_coef_experimental_data(small_tables_format).xlsx"
    from decimal import Decimal
    # Writing the values
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    table_betas = np.reshape(np.array(df['beta[deg]']), (5, 6), order='F')
    table_thetas = np.reshape(np.array(df['theta[deg]']), (5, 6), order='F')
    table_angle_pairs = [['(' + str('%.2f' % table_betas[i, j]) + ', ' + str('%.2f' % table_thetas[i, j]) + ')' for j in range(6)] for i in range(5)]  # todo: CHANGE THETA ORDER TO MATCH 2 PLOTS
    pd.DataFrame(table_angle_pairs).to_excel(writer, header=False, index=False, sheet_name='angle_pairs')
    table_Ci_num = []
    table_Ci_str = []
    for idx_str in ['Cx_Ls', 'Cy_Ls', 'Cz_Ls', 'Cxx_Ls', 'Cyy_Ls', 'Czz_Ls']:
        table_Ci_num.append(np.reshape(np.array(df[idx_str]), (5, 6), order='F'))
        table_Ci_str.append(np.reshape(['{:.2E}'.format(Decimal(x)) for x in np.array(df[idx_str])], (5, 6), order='F'))
        pd.DataFrame(table_Ci_str[-1]).to_excel(writer, header=False, index=False, sheet_name=idx_str)
    writer.close()

    # Color Fill each cell of each table in the Excel file
    import matplotlib.pyplot as plt
    import matplotlib
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font

    C_Ci_grid_flat_Ls = aero_coef(xx.flatten(), yy.flatten(), method=method, coor_system='Ls')

    wb = load_workbook(filename=filename)
    for i, idx_str in enumerate(['Cx_Ls', 'Cy_Ls', 'Cz_Ls', 'Cxx_Ls', 'Cyy_Ls', 'Czz_Ls']):
        ws = wb[idx_str]

        if color_interval == 'same_as_zoomin_plot':
            absmax = max(abs(C_Ci_grid_flat_Ls[i]))
            normal_min = -absmax
            normal_max = absmax
        else:
            normal_min = np.min(table_Ci_num[i])
            normal_max = np.max(table_Ci_num[i])
        normal = (table_Ci_num[i] - normal_min) / (normal_max - normal_min)
        for j in range(5):
            for k in range(6):
                color_rgb_code = matplotlib.colors.to_hex(plt.cm.Spectral_r(normal)[j, k])[1:]
                wc = ws.cell(row=j + 1, column=k + 1)
                if color_in=='cell':
                    wc.fill = PatternFill(start_color=color_rgb_code, end_color=color_rgb_code, fill_type='solid')
                elif color_in=='font':
                    wc.font = Font(color=color_rgb_code)
    wb.save(filename=filename)
    wb.close()
    # # If a table is desired in Python, then:
    # for i, idx_str in enumerate(['Cx_Ls']):
    #     plt.figure()
    #     normal = (table_Ci_num[i] - np.min(table_Ci_num[i])) / (np.max(table_Ci_num[i]) - np.min(table_Ci_num[i]))
    #     plt.table(table_Ci_str[i], loc='center', cellColours=plt.cm.Spectral_r(normal))
color_table_SOH_values(color_in='cell')

def colormap_2D_vs_Cosine():
    beta_angle_step = 0.1 # in degrees.
    theta_angle_step = 0.2 # in degrees.
    betas = np.arange(rad(0), rad(90), rad(beta_angle_step))
    thetas = np.arange(rad(-89.9), rad(89.9) + rad(theta_angle_step) * 0.012345, rad(theta_angle_step))
    xx, yy = np.meshgrid(betas, thetas)
    from buffeting import U_bar_func
    from straight_bridge_geometry import g_node_coor
    def truncate_colormap(cmap, minval=0.0, maxval=1.0, n=100):
        new_cmap = matplotlib.colors.LinearSegmentedColormap.from_list(
            'trunc({n},{a:.2f},{b:.2f})'.format(n=cmap.name, a=minval, b=maxval),
            cmap(np.linspace(minval, maxval, n)))
        return new_cmap
    U = U_bar_func(g_node_coor)[0]
    Uyz = U * np.sqrt(1 - np.sin(xx)**2 * np.cos(yy)**2)
    Ucosine = U * np.cos(xx)
    # zz_1 = Uyz - Ucosine
    zz_1 = (Uyz - Ucosine)/Uyz * 100
    theta_yz = np.arcsin(np.sin(yy) / np.sqrt(1 - np.sin(xx)**2 * np.cos(yy)**2))
    zz_2 = deg(theta_yz - yy)  # this is to force an extra tick at the colorbar at 90deg
    # zz_2 = (theta_yz - yy)/theta_yz * 100  # this is to force an extra tick at the colorbar at 90deg
    for i,zz in enumerate([zz_1, zz_2]):
        # Finding the coefficient of determination R_squared
        fig = plt.figure(figsize=(5, 4), dpi=300)
        ax = plt.axes()
        titles = [r'$(U_{yz} - U\/\cos(\beta))$ / $U_{yz}$ [%]', r'$\theta_{yz}-\theta$ [$\degree$]']
        # titles = [r'$(U_{yz} - U\/\cos(\beta))$ / $U_{yz}$ [%]', r'$(\theta_{yz}-\theta)$ / $\theta_{yz}$ [%]']
        str_save = ['Uyz-U', 'thetayz-theta']
        clims = [(0, 100), (-90, 90)]
        # clims = [(0, 100), (0, 100)]
        plt.title(titles[i])
        cmap = plt.get_cmap(matplotlib.cm.Spectral_r)
        if i == 0:
        # if i in [0,1]:
            cmap = truncate_colormap(cmap, minval=0.5)
        cmap_norm = matplotlib.colors.Normalize(vmin=clims[i][0], vmax=clims[i][1])
        scalarMap = matplotlib.cm.ScalarMappable(norm=cmap_norm, cmap=cmap)
        sct = ax.scatter(deg(xx.flatten()), deg(yy.flatten()), s = np.ones(len(zz.flatten()))*10, alpha=1,marker="o", c=scalarMap.to_rgba(zz.flatten()), vmin=clims[i][0], vmax=clims[i][1])
        sct.set_clim(0, 15)
        cbar = fig.colorbar(matplotlib.cm.ScalarMappable(norm=cmap_norm, cmap=cmap), ax=ax, alpha=1)
        if i == 1:
            cbar.set_ticks([-90,-60,-30,0,30,60,90])
        ax.set_xlabel(r'$\beta\/[\degree]$')
        ax.set_ylabel(r'$\theta\/[\degree]$')
        ax.set_xlim(deg(min(betas)), deg(max(betas)))
        plt.xticks(np.arange(0, 91, 15))
        plt.yticks(np.arange(-90, 91, 15))
        ax.set_ylim(-90, 90)
        plt.tight_layout()
        plt.savefig(r'aerodynamic_coefficients/plots/'+str_save[i]+'.png')
        plt.close()
colormap_2D_vs_Cosine()

#####################################################################################################################
# Looking at the derivatives
#####################################################################################################################
# beta_angle_step = 1  # in degrees.
# betas = np.arange(rad(-179), rad(180)+rad(beta_angle_step)*0.012345, rad(beta_angle_step))
# C_Ci_fit_at_SOH_cons = aero_coef(betas, np.zeros(betas.shape), method='2D_fit_cons', coor_system='Ls')
# C_Ci_fit_at_SOH_free = aero_coef(betas, np.zeros(betas.shape), method='2D_fit_free', coor_system='Ls')
# C_Ci_fit_at_SOH_derivatives_cons = aero_coef_derivatives(betas, np.zeros(betas.shape), method='2D_fit_cons', coor_system='Ls')
# C_Ci_fit_at_SOH_derivatives_free = aero_coef_derivatives(betas, np.zeros(betas.shape), method='2D_fit_free', coor_system='Ls')
# dof_idx = 1
# plt.figure()
# plt.title('cons')
# plt.plot(deg(betas), C_Ci_fit_at_SOH_cons[dof_idx], label='cons')
# plt.plot(deg(betas), C_Ci_fit_at_SOH_derivatives_cons[0,dof_idx], label='db cons')
# plt.plot(deg(betas), C_Ci_fit_at_SOH_derivatives_cons[1,dof_idx], label='dt cons')
# plt.legend()
# plt.figure()
# plt.title('free')
# plt.plot(deg(betas), C_Ci_fit_at_SOH_free[dof_idx], label='free')
# plt.plot(deg(betas), C_Ci_fit_at_SOH_derivatives_free[0,dof_idx], label='db free')
# plt.plot(deg(betas), C_Ci_fit_at_SOH_derivatives_free[1,dof_idx], label='dt free')
# plt.legend()



